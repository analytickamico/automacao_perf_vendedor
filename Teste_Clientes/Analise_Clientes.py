import streamlit as st
import pandas as pd
import pydeck as pdk
import plotly.express as px
import numpy as np
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import io


def load_data():
    df = pd.read_csv('customer_data.csv')
    df['maior_mes'] = pd.to_datetime(df['maior_mes'])
    df['mes_ano'] = df['maior_mes'].dt.strftime('%Y-%m')
    df['latitude'] = pd.to_numeric(df['latitude'], errors='coerce')
    df['longitude'] = pd.to_numeric(df['longitude'], errors='coerce')
    df['Canal_Venda'] = df['Canal_Venda'].fillna('Not Specified')
    df['equipe'] = df['equipe'].fillna('Not Assigned')
    df['Vendedor'] = df['Vendedor'].fillna('Not Assigned')
    return df

def generate_route_suggestion(df):
    """Gera sugestão de rota priorizando clientes por recência"""
    routes = {}
    
    for vendedor in df['Vendedor'].unique():
        vendor_df = df[df['Vendedor'] == vendedor].copy()
        
        if len(vendor_df) > 0:
            # Ordenar por recência (prioridade) e depois por proximidade (cidade/bairro)
            vendor_df = vendor_df.sort_values(
                ['Recencia', 'cidade', 'bairro'],
                ascending=[True, True, True]  # True para recência para priorizar menores valores
            )
            
            routes[vendedor] = vendor_df[['Cliente', 'endereco', 'maior_mes', 'Recencia']]
    
    return routes

def generate_pdf_report(df, routes):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus.flowables import KeepTogether
    import io
    
    buffer = io.BytesIO()
    page_width, page_height = landscape(A4)
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=0.5*inch,
        rightMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Título principal
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        spaceAfter=20
    )
    story.append(Paragraph('Relatório de Clientes e Rotas Sugeridas', title_style))
    
    # Lista de clientes por vendedor
    for vendedor in sorted(df['Vendedor'].unique()):
        vendor_df = df[df['Vendedor'] == vendedor].sort_values('maior_mes', ascending=True)
        
        if len(vendor_df) > 0:
            # Título do vendedor
            story.append(Paragraph(f'Vendedor: {vendedor}', styles['Heading2']))
            
            # Tabela de clientes
            df_display = vendor_df[[
                'maior_mes', 'Cliente', 'endereco',
                'marcas_concatenadas', 'Ticket_Medio'
            ]].copy()
            
            df_display['maior_mes'] = pd.to_datetime(df_display['maior_mes']).dt.strftime('%d/%m/%Y')
            df_display['Ticket_Medio'] = df_display['Ticket_Medio'].apply(lambda x: f'R$ {x:,.2f}')
            
            data = [['Última Compra', 'Cliente', 'Endereço', 'Marcas', 'Ticket Médio']]
            data.extend(df_display.values.tolist())
            
            # Ajustar larguras das colunas
            col_widths = [1*inch, 2*inch, 3.5*inch, 3*inch, 1.2*inch]
            table = Table(data, colWidths=col_widths, repeatRows=1)
            
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
            
            # Rota sugerida
            story.append(Paragraph('Rota Sugerida de Visitas', styles['Heading3']))
            
            route_data = [['Sequência', 'Cliente', 'Endereço', 'Última Visita', 'Recência']]
            vendor_route = vendor_df.sort_values(['Recencia', 'maior_mes'])
            
            for i, (_, row) in enumerate(vendor_route.iterrows(), 1):
                route_data.append([
                    str(i),
                    row['Cliente'],
                    row['endereco'],
                    pd.to_datetime(row['maior_mes']).strftime('%d/%m/%Y'),
                    f"{int(row['Recencia'])} meses"
                ])
            
            route_table = Table(route_data, colWidths=[0.7*inch, 2*inch, 3.5*inch, 1*inch, 1*inch])
            route_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(route_table)
            story.append(PageBreak())
    
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_excel_report(df):
    import pandas as pd
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        
        for vendedor in sorted(df['Vendedor'].unique()):
            vendor_df = df[df['Vendedor'] == vendedor].copy()
            
            # Aba Lista de Clientes
            sheet_name = f"{vendedor[:30]}_Clientes"
            
            clientes_df = vendor_df[[
                'maior_mes', 'Cliente', 'endereco',
                'marcas_concatenadas', 'Ticket_Medio'
            ]].sort_values('maior_mes', ascending=True)
            
            clientes_df['maior_mes'] = pd.to_datetime(clientes_df['maior_mes']).dt.strftime('%d/%m/%Y')
            clientes_df.columns = ['Última Compra', 'Cliente', 'Endereço', 'Marcas', 'Ticket Médio']
            
            clientes_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Formatar planilha de clientes
            ws = workbook[sheet_name]
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrapText=True, vertical='center')
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
            
            ws.column_dimensions['A'].width = 12  # Última Compra
            ws.column_dimensions['B'].width = 30  # Cliente
            ws.column_dimensions['C'].width = 50  # Endereço
            ws.column_dimensions['D'].width = 40  # Marcas
            ws.column_dimensions['E'].width = 15  # Ticket Médio
            
            for cell in ws['E']:
                if cell.row > 1:
                    cell.number_format = 'R$ #,##0.00'
            
            # Aba Rota Sugerida
            sheet_name = f"{vendedor[:30]}_Rota"
            
            # Ordenar por recência (decrescente) e depois por data da última compra
            rota_df = vendor_df[[
                'Cliente', 'endereco', 'maior_mes', 'Recencia'
            ]].sort_values(['Recencia', 'maior_mes'], ascending=[False, True])
            
            rota_df['Sequência'] = range(1, len(rota_df) + 1)
            rota_df['maior_mes'] = pd.to_datetime(rota_df['maior_mes']).dt.strftime('%d/%m/%Y')
            rota_df['Recencia'] = rota_df['Recencia'].apply(lambda x: f'{int(x)} meses')
            
            rota_df = rota_df[['Sequência', 'Cliente', 'endereco', 'maior_mes', 'Recencia']]
            rota_df.columns = ['Sequência', 'Cliente', 'Endereço', 'Última Visita', 'Recência']
            
            rota_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Formatar planilha de rota
            ws = workbook[sheet_name]
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrapText=True, vertical='center')
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
            
            ws.column_dimensions['A'].width = 10  # Sequência
            ws.column_dimensions['B'].width = 30  # Cliente
            ws.column_dimensions['C'].width = 50  # Endereço
            ws.column_dimensions['D'].width = 15  # Última Visita
            ws.column_dimensions['E'].width = 12  # Recência
    
    output.seek(0)
    return output

def create_map(df_filtered, team=None):
    if df_filtered.empty:
        return None

    map_data = df_filtered[df_filtered['Recencia'] <= 4].dropna(subset=['latitude', 'longitude']).copy()

    if len(map_data) == 0:
        return None
    # State boundaries layer
    states_layer = pdk.Layer(
        'GeoJsonLayer',
        'https://raw.githubusercontent.com/tbrugz/geodata-br/master/geojson/geojs-100-mun.json',
        stroked=True,
        filled=False,
        get_line_color=[169, 169, 169],
        line_width_min_pixels=1,
        pickable=False
    )

    # Cities boundaries layer
    cities_layer = pdk.Layer(
        'GeoJsonLayer',
        'https://raw.githubusercontent.com/tbrugz/geodata-br/master/geojson/geojs-35-mun.json',
        stroked=True,
        filled=False,
        get_line_color=[211, 211, 211],
        line_width_min_pixels=0.5,
        pickable=False
    )

    def get_color_by_recency(recency):
        colors = {
            0: [0, 0, 255, 180],
            1: [0, 255, 0, 180],
            2: [255, 255, 0, 180],
            3: [255, 165, 0, 180],
            4: [255, 0, 0, 180]
        }
        return colors[min(int(recency), 4)]

    map_data['color'] = map_data['Recencia'].astype(float).apply(get_color_by_recency)

    point_layer = pdk.Layer(
        'ScatterplotLayer',
        data=map_data,
        get_position=['longitude', 'latitude'],
        get_fill_color='color',
        get_radius=300,
        pickable=True,
        opacity=0.8,
        stroked=True,
        filled=True,
        radius_min_pixels=5,
        radius_max_pixels=15,
        line_width_min_pixels=1
    )

    try:
        center_lat = map_data['latitude'].mean()
        center_lon = map_data['longitude'].mean()
    except:
        center_lat = -23.5505
        center_lon = -46.6333

    view_state = pdk.ViewState(
        latitude=center_lat,
        longitude=center_lon,
        zoom=9,
        pitch=0
    )

    return pdk.Deck(
        layers=[states_layer, cities_layer, point_layer],
        initial_view_state=view_state,
        tooltip={
            "html": "<b>{Cliente}</b><br/>"
                   "Cidade: {cidade}<br/>"
                   "Bairro: {bairro}<br/>"
                   "Recência: {Recencia} meses<br/>"
                   "Marcas: {marcas_concatenadas}",
        },
        map_style="light"
    )
def generate_sales_report(df, filename='sales_report.pdf'):
    doc = SimpleDocTemplate(filename, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30
    )
    story.append(Paragraph('Análise de Clientes - Relatório Regional', title_style))
    story.append(Spacer(1, 12))

    # Recency Distribution
    story.append(Paragraph('Distribuição por Recência', styles['Heading2']))
    recency_data = df[df['Recencia'] <= 4].groupby('Recencia').size()
    recency_table = Table([
        ['Período', 'Quantidade', '% Total'],
        *[
            [f'{i} meses', count, f'{count/len(df)*100:.1f}%']
            for i, count in recency_data.items()
        ]
    ])
    recency_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(recency_table)
    story.append(Spacer(1, 20))

    # Location stats
    story.append(Paragraph('Concentração Regional', styles['Heading2']))
    location_stats = (df.groupby(['cidade', 'bairro'])
                     .agg({
                         'Cliente': 'count',
                         'Ticket_Medio': 'mean',
                         'Recencia': 'mean'
                     })
                     .sort_values('Cliente', ascending=False)
                     .head(10))
    
    location_data = [['Região', 'Clientes', 'Ticket Médio', 'Recência']]
    for (city, nbhd), row in location_stats.iterrows():
        location_data.append([
            f'{nbhd} ({city})',
            str(int(row['Cliente'])),
            f'R$ {row["Ticket_Medio"]:.2f}',
            f'{row["Recencia"]:.1f}'
        ])
    
    location_table = Table(location_data)
    location_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(location_table)

    doc.build(story)

def generate_enhanced_report(df, filename='customer_report.pdf'):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    import io
    import matplotlib.pyplot as plt
    
    doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1
    )
    story.append(Paragraph('Análise de Clientes', title_style))
    
    # Distribuição por Recência (Gráfico)
    recency_data = df[df['Recencia'] <= 5].groupby('Recencia')['Cliente'].count()
    plt.figure(figsize=(8, 4))
    recency_data.plot(kind='bar')
    plt.title('Distribuição por Recência')
    plt.xlabel('Meses')
    plt.ylabel('Número de Clientes')
    
    img_buf = io.BytesIO()
    plt.savefig(img_buf, format='png')
    img_buf.seek(0)
    story.append(Image(img_buf, width=400, height=200))
    plt.close()
    
    story.append(Spacer(1, 20))
    
    # Métricas Gerais
    metrics_data = [
        ['Indicador', 'Valor'],
        ['Total de Clientes', len(df)],
        ['Ticket Médio', f'R$ {df["Ticket_Medio"].mean():.2f}'],
        ['Média de Positivações', f'{df["Positivacao"].mean():.1f}'],
        ['Clientes Inadimplentes', f'{len(df[df["status_inadimplente"]=="Inadimplente"])} ({len(df[df["status_inadimplente"]=="Inadimplente"])/len(df)*100:.1f}%)']
    ]
    
    metrics_table = Table(metrics_data, colWidths=[200, 200])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 20))
    
    # Análise por UF
    uf_analysis = df.groupby('uf').agg({
        'Cliente': 'count',
        'Ticket_Medio': 'mean',
        'Positivacao': 'mean',
        'status_inadimplente': lambda x: (x == 'Inadimplente').mean() * 100
    }).reset_index()
    
    uf_data = [['UF', 'Clientes', 'Ticket Médio', 'Positivações', '% Inadimplência']]
    for _, row in uf_analysis.iterrows():
        uf_data.append([
            row['uf'],
            str(row['Cliente']),
            f'R$ {row["Ticket_Medio"]:.2f}',
            f'{row["Positivacao"]:.1f}',
            f'{row["status_inadimplente"]:.1f}%'
        ])
    
    uf_table = Table(uf_data, colWidths=[80, 80, 100, 80, 100])
    uf_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(Paragraph('Análise por UF', styles['Heading2']))
    story.append(uf_table)
    story.append(Spacer(1, 20))
    
    # Clientes Inadimplentes
    if len(df[df['status_inadimplente'] == 'Inadimplente']) > 0:
        defaulters = df[df['status_inadimplente'] == 'Inadimplente'].copy()
        defaulters_data = [['Cliente', 'Cidade/UF', 'Valor', 'Títulos', 'Último Pedido']]
        
        for _, row in defaulters.iterrows():
            defaulters_data.append([
                row['Cliente'],
                f"{row['cidade']}/{row['uf']}",
                f"R$ {row['vlr_inadimplente']:.2f}",
                str(row['qtd_titulos']),
                row['mes_ano']
            ])
        
        defaulters_table = Table(defaulters_data, colWidths=[150, 100, 100, 80, 100])
        defaulters_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(Paragraph('Clientes Inadimplentes', styles['Heading2']))
        story.append(defaulters_table)
    
    # Build PDF
    doc.build(story)

def main():
    st.set_page_config(layout="wide")
    
    # Carregar dados
    df = load_data()
    
    
    # Filtros no sidebar
    with st.sidebar:
        st.markdown("## Filtros")
        canal_venda = st.multiselect('Canal de Venda', options=sorted(df['Canal_Venda'].unique()))
        equipe = st.multiselect('Equipe', options=sorted(df['equipe'].unique()))
        
        if equipe:
            vendedores_disponiveis = sorted(df[df['equipe'].isin(equipe)]['Vendedor'].unique())
        else:
            vendedores_disponiveis = sorted(df['Vendedor'].unique())
        
        vendedor = st.multiselect('Vendedor', options=vendedores_disponiveis)
        
        recencia_range = st.slider('Recência (meses)', 
                                 min_value=0, 
                                 max_value=12,
                                 value=(0, 4))
    
    # Aplicar filtros
    filtered_df = df.copy()
    if canal_venda:
        filtered_df = filtered_df[filtered_df['Canal_Venda'].isin(canal_venda)]
    if equipe:
        filtered_df = filtered_df[filtered_df['equipe'].isin(equipe)]
    if vendedor:
        filtered_df = filtered_df[filtered_df['Vendedor'].isin(vendedor)]
    
    filtered_df = filtered_df[
        (filtered_df['Recencia'] >= recencia_range[0]) & 
        (filtered_df['Recencia'] <= recencia_range[1])
    ]

    # Mapa
    st.markdown("## Mapa de Clientes")
    map_chart = create_map(filtered_df)
    if map_chart:
        st.pydeck_chart(map_chart)
    else:
        st.warning("Não foi possível criar o mapa. Verifique se existem dados com coordenadas válidas.")

    # Distribuição de Clientes
    st.markdown("## Distribuição de Clientes")
    
    col1, col2 = st.columns([3, 2])
    
    with col1:
        # Gráfico de barras
        recency_data = filtered_df.groupby('Recencia')['Cliente'].count().reset_index()
        fig = px.bar(recency_data, x='Recencia', y='Cliente',
                    labels={'Cliente': 'Quantidade', 'Recencia': 'Meses'})
        
        fig.update_layout(
            showlegend=False,
            yaxis_showticklabels=False,
            xaxis_tickmode='linear',
            xaxis_dtick=1
        )
        
        fig.update_traces(
            text=recency_data['Cliente'],
            textposition='outside',
            texttemplate='%{text:,}'
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    with col2:
        # Tabela de distribuição
        total_clientes = len(filtered_df)
        recency_table = (filtered_df.groupby('Recencia')
                        .agg({'Cliente': 'count'})
                        .reset_index()
                        .sort_values('Recencia'))
        
        # Calcular percentuais
        recency_table['% Total'] = (recency_table['Cliente'] / total_clientes * 100).round(2)
        
        # Formatar colunas
        recency_table['Período'] = recency_table['Recencia'].apply(lambda x: f'{int(x)} meses')
        recency_table['% Total'] = recency_table['% Total'].apply(lambda x: f'{x:.2f}%')
        recency_table = recency_table.rename(columns={'Cliente': 'Quantidade'})
        
        # Adicionar total
        total_row = pd.DataFrame({
            'Período': ['Total'],
            'Quantidade': [total_clientes],
            '% Total': ['100.00%']
        })
        
        # Selecionar e ordenar colunas
        final_table = pd.concat([
            recency_table[['Período', 'Quantidade', '% Total']], 
            total_row
        ], ignore_index=True)
        
        st.dataframe(
            final_table,
            hide_index=True,
            use_container_width=True
        )

    st.markdown("## Lista de Clientes")
    
    # Métricas
    total_monetario = filtered_df['Monetario'].sum()
    positivacao=filtered_df['Positivacao'].sum()

    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Total de Clientes", f"{total_clientes:,}")
    with col2:
        st.metric(
            "Ticket Médio",
            f"R$ {(total_monetario/positivacao if total_clientes > 0 else 0):,.2f}"
        )

    # Botão de relatório
    if st.button('Gerar Relatório Excel'):
        excel_buffer = generate_excel_report(filtered_df)
        st.download_button(
            label="Baixar Relatório Excel",
            data=excel_buffer,
            file_name="relatorio_clientes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    # Lista de clientes
    display_df = filtered_df[[
        'Vendedor', 'Canal_Venda', 'Cliente', 'maior_mes', 'endereco',
        'marcas_concatenadas', 'Ticket_Medio', 'Positivacao', 'Monetario'
    ]].copy()
    
    display_df['mes_ano'] = pd.to_datetime(display_df['maior_mes']).dt.strftime('%m-%Y')
    display_df['Ticket_Medio'] = display_df['Ticket_Medio'].map('R$ {:,.2f}'.format)
    display_df['Monetario'] = display_df['Monetario'].map('R$ {:,.2f}'.format)
    
    display_cols = {
        'Vendedor': 'Vendedor',
        'Canal_Venda': 'Canal',
        'Cliente': 'Cliente',
        'mes_ano': 'Última Compra',
        'endereco': 'Endereço',
        'marcas_concatenadas': 'Marcas',
        'Ticket_Medio': 'Ticket Médio',
        'Positivacao': 'Positivações',
        'Monetario': 'Monetário'
    }
    
    st.dataframe(
        display_df.sort_values('maior_mes', ascending=True)[display_cols.keys()].rename(columns=display_cols),
        hide_index=True,
        use_container_width=True
    )

if __name__ == '__main__':
    main()